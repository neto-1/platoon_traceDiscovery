from __future__ import print_function
from sklearn import tree
import numpy
import numpy as np
from scipy.optimize import minimize
#from scipy.optimize import minimize_scalar
from gurobipy import *
import string
from neo4jrestclient.client import GraphDatabase
from scipy import optimize
import subprocess
import json
import platform
import yaml
import psycopg2
import pycurl
import collections
import pycurl
from StringIO import StringIO
import time
from database import Experiment
from gurobipy import *




src = "C:/Users/gerrit/Documents/platooning/src"


def get_process_return_value(process, verbose=False):
    """
    Get the return value of process (i.e. the last print statement)
    :param process: The process returned from subprocess.popen
    :param verbose: Whether or not the output of the process should be printed
    """
    output = None

    # poll will return the exit code if the process is completed otherwise it returns null
    while process.poll() is None:
        line = process.stdout.readline()
        if not line:
            break
        output = line  # last print statement
        if verbose:
            print(line.rstrip().decode('utf-8'))
    return json.loads(output)  # parse JSON



def length_of_platoon_dist(start_veh1, start_veh2, end_veh1, end_veh2, weight_nu):
    # define function
    # geometric_median_function = lambda x: (np.sqrt((start_veh1[0] - x[0])**2 + (start_veh1[1] - x[1])**2) + np.sqrt((start_veh2[0] - x[0])**2 + (start_veh2[1] - x[1])**2) + np.sqrt((end_veh1[0] - x[2])**2 + (end_veh1[1] - x[3])**2) + np.sqrt((end_veh2[0] - x[2])**2 + (end_veh2[1] - x[3])**2) + 2*np.sqrt((x[0] - x[2])**2 + (x[1] - x[3])**2) - (np.sqrt((x[0] - x[2])**2 + (x[1] - x[3])**2)) * weight_nu)

    geometric_median_function =(-np.exp(-(start_veh1[0] - x[0]) ** 2 + -(start_veh1[1] - x[1]) ** 2) + -np.exp(-(start_veh2[0] - x[0]) ** 2 + -(start_veh2[1] - x[1]) ** 2) + -np.exp(-(end_veh1[0] - x[2]) ** 2 + -(end_veh1[1] - x[3]) ** 2) + -np.exp(-(end_veh2[0] - x[2]) ** 2 + -(end_veh2[1] - x[3]) ** 2) + 2 * -np.exp(-(x[0] - x[2]) ** 2 + -(x[1] - x[3]) ** 2) - (-np.exp(-(x[0] - x[2]) ** 2 + -(x[1] - x[3]) ** 2)) * weight_nu)
    # (((start_veh1[0] - x[0])**2 + (start_veh1[1] - x[1])**2)**0.5 + ((start_veh2[0] - x[0])**2 + (start_veh2[1] - x[1])**2)**0.5 + ((end_veh1[0] - x[2])**2 + (end_veh1[1] - x[3])**2)**0.5 + ((end_veh2[0] - x[2])**2 + (end_veh2[1] - x[3])**2)**0.5) + 2*((x[0] - x[2])**2 + (x[1] - x[3])**2)**0.5 - (((x[0] - x[2])**2 + (x[1] - x[3])**2)**0.5) * weight_nu
    # calculate air distance of compared vehicles
    platoon_distance = optimize.minimize_scalar(geometric_median_function, [50, 50, 50, 50], method='Brent')
    # print('plat dist: ' + str(platoon_distance.fun))
    return platoon_distance


def air_dist(start_veh1, start_veh2, end_veh1, end_veh2):
    v1 = ((start_veh1[0] - end_veh1[0]) ** 2 + (start_veh1[1] - end_veh1[1]) ** 2) ** 0.5
    v2 = ((start_veh2[0] - end_veh2[0]) ** 2 + (start_veh2[1] - end_veh2[1]) ** 2) ** 0.5
    distance = v1 + v2

    return distance


def get_nu(p12, p13, p22, p23, weight):

    # calculate platooning distance
    platooning_distance = length_of_platoon_dist(p12, p13, p22, p23, weight)

    # calculate air distance
    air_distance = air_dist(start_veh1, start_veh2, end_veh1, end_veh2)

    # try to adjust the linear weight so that airdistance and platooning distance are nearly the same
    while abs(air_distance - platooning_distance.fun) > 0.01:

        if air_distance < platooning_distance.fun:
            weight += 0.1
        else:
            weight -= 0.001
        platooning_distance = length_of_platoon_dist(p12, p13, p22, p23, weight)
        print('air dist: ' + str(air_distance) + ' plat dist: ' + str(platooning_distance.fun) + ' with weight of: ' + str(weight) + ' with points: ' + str(platooning_distance.x))
        # print(weight)
    # return the weight of nu
    return weight

# starting point of veh1
start_veh1 = np.array([1, 1])
# starting point of veh2
start_veh2 = np.array([2, 1])
# end point ofveh1
end_veh1 = np.array([1, 50])
# end point of veh2
end_veh2 = np.array([2, 50])
linear_weight = 0

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

#ws = wb.active
#wb = load_workbook(filename = 'BerechnungETA.xlsx')
#sheet_ranges = wb['range names']

wb = load_workbook("C:\\Users\\gerrit\\Documents\\platooning\\src\\group\\params\\BerechnungETA.xlsx")
#print(wb)
ws = wb['Tabelle1']
#print(ws)
eta_values = {}
der_name_der_spalte = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']

linear_weight = 2
#print(ws['B2'].value)
for i in range(2, 10, 1):
    start_veh1 = np.array([ws[str(der_name_der_spalte[0] + str(i))].value, ws[str(der_name_der_spalte[1] + str(i))].value])
    start_veh2 = np.array([ws[str(der_name_der_spalte[2] + str(i))].value, ws[str(der_name_der_spalte[3] + str(i))].value])
    end_veh1 = np.array([ws[str(der_name_der_spalte[4] + str(i))].value, ws[str(der_name_der_spalte[5] + str(i))].value])
    end_veh2 = np.array([ws[str(der_name_der_spalte[6] + str(i))].value, ws[str(der_name_der_spalte[7] + str(i))].value])
    eta_value = {}

    eta_value['air_dist'] = []
    eta_value['air_dist'].append(0)
    eta_value['platoon_dist'] = []
    eta_value['platoon_dist'].append(5)

    # start_veh1 = np.array([1, 1])
    # # starting point of veh2
    # start_veh2 = np.array([2, 1])
    # # end point ofveh1
    # end_veh1 = np.array([1, 50])
    # # end point of veh2
    # end_veh2 = np.array([2, 50])
    linear_weight = 1
    while eta_value['air_dist'][0] - eta_value['platoon_dist'][0] > 0.05 or eta_value['platoon_dist'][0] - eta_value['air_dist'][0] > 0.05:

        cmd = "Rscript " + "GeometricApproachETA.R" + " " + src + " " + str(start_veh1[0]) + " " + str(
            start_veh1[1]) + " " + str(start_veh2[0]) + " " + str(start_veh2[1]) + " " + str(end_veh1[0]) + " " + str(
            end_veh1[1]) + " " + str(end_veh2[0]) + " " + str(end_veh2[1]) + " " + str(linear_weight)
        return_eta = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE)
        eta_value = get_process_return_value(return_eta)
        return_eta.wait()

        # operate the function
        if eta_value['air_dist'][0] < eta_value['platoon_dist'][0]:
            linear_weight += 0.1
        else:
            linear_weight -= 0.001
        print(linear_weight)
        print("---")
        # print('weight ' + str(linear_weight))
        #print(eta_value['air_dist'])
        #print(eta_value['platoon_dist'])
        print(eta_value['air'])
        print(eta_value['platoon_dist'])
    print('done')
    eta_values['line_' + str(i)] = [eta_value['air_dist'][0], eta_value['platoon_dist'][0], linear_weight]
# the value of nu strich
#print(res)

print(eta_values)
