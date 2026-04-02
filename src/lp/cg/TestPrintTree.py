from gurobipy import *
import networkx as nx
from neo4jrestclient.client import GraphDatabase
from neo4jrestclient import client
import time
import lppl
import numpy as np
import createInitPaths
import lpMaster
#import lpSub
import lpSubDual

from vehicle import Vehicle
from route import Route
### initialization of graph database neo4j
db = GraphDatabase("http://localhost:7474", username="neo4j", password="12345678")
# db = GraphDatabase("http://localhost:7474", username="neo4j", password="1234567890")
import random


##### first get the set of vehicles to get all vehicles considered (not only the groups)
query = "MATCH (n:PlatooningUtility) RETURN n.vector_degree as richtung, n.gradient as gradient, n.overlapVol as overlap, n.overlapArea as area"
results = db.query(query)

#####Erstelle Feature Liste aus results ([[Wert1,Wert2,Wert3,Wert4],[...]]
#####Erstelle Labels Liste mit [0,1] unter if Schleife mit x (hier noch Random)
features=[]
labels=[]
for i in results:
    features.append(i)
    x= (i[0]+i[1]+(i[2]+i[3]))
    print(x)
    if x<=1.00:
        x1=1
    if x>1.00:
        x1=2
    if x>1.20:
        x1=3
    if x>1.60:
        x1=4
    if x>1.85:
        x1=5

    labels.append(x1)
print(labels)
print("-----------")


##################################################Decisiontree#####################################################################
from sklearn import tree
#beide Test Listen erstellen
features_test = features
labels_test = labels
#classifier, Methode "wie" die Daten "abgeschnitten" werden hier
clf = tree.DecisionTreeClassifier()
clf = clf.fit(features_test, labels_test)
print(clf.fit)
#predict= "Vorhersage-Wert"
clf.predict([[0.986979408351784, 0.683106271366098, 0.134448467414797, 0.325965734398636]])
clf.predict_proba([[0.986979408351784, 0.683106271366098, 0.134448467414797, 0.325965734398636]])

#package zum exportieren des Dec Tree
import graphviz
dot_data = tree.export_graphviz(clf, out_file=None)
graph = graphviz.Source(dot_data)
#Lade txt Datei in http://www.webgraphviz.com/
tree.export_graphviz(clf, out_file='tree3.txt')
print("done")
##################################################Decisiontree#####################################################################





#
# X = [[0, 0], [1, 1]]
# Y = [0, 1]
# clf = tree.DecisionTreeClassifier()
# clf = clf.fit(X, Y)
#
# prediction = clf.predict([[2., 2.]])
#
# print(prediction)
#
#
#
# from sklearn.datasets import load_iris
# from sklearn import tree
# iris = load_iris()
# clf = tree.DecisionTreeClassifier()
# clf = clf.fit(iris.data, iris.target)
#
# print(iris.target)
#
#
# import graphviz
# dot_data = tree.export_graphviz(clf, out_file=None)
# graph = graphviz.Source(dot_data)
# graph.render("iris")



# dot_data = tree.export_graphviz(clf, out_file=None,
#                          feature_names=iris.feature_names,
#                          class_names=iris.target_names,
#                          filled=True, rounded=True,
#                          special_characters=True)
# graph = graphviz.Source(dot_data)

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

# ws = wb.active
# wb = load_workbook(filename = 'BerechnungETA.xlsx')
# sheet_ranges = wb['range names']

wb = load_workbook("C:\\Users\\gerrit\\Documents\\platooning\\src\\lp\\cg\\Lambdas_SP_Savings.xlsx")
# print(wb)
ws = wb['Tabelle1']
# print(ws)

der_name_der_spalte = ['B', 'C', 'D', 'E', 'F']


#print(ws['B2'].value)
for i in range(2, 77, 1):
    Datensatz_T = np.array([ws[str(der_name_der_spalte[0] + str(i))].value, ws[str(der_name_der_spalte[1] + str(i))].value, ws[str(der_name_der_spalte[2] + str(i))].value, ws[str(der_name_der_spalte[3] + str(i))].value, ws[str(der_name_der_spalte[4] + str(i))].value])
    print(Datensatz_T)

